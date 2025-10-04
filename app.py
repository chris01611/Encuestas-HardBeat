
from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def survey():
    if request.method == 'POST':
        rating = request.form.get('rating')
        comments = request.form.get('comments')
        sucursal = request.args.get('sucursal', 'Desconocida')
        bano = request.args.get('bano', 'Desconocido')

        file_name = 'respuestas_encuesta.xlsx'
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Respuestas"
            ws.append(["Sucursal", "Baño", "Calificación", "Comentarios"])

        ws.append([sucursal, bano, rating, comments])

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_name)
        return redirect('/')

    return render_template("formulario_encuesta.html")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
